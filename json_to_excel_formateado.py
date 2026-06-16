#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
JSON to Excel Converter with Professional Formatting
Converts resultados_validados_gemini.json to Excel with:
- Table formatting
- Auto-fitted columns
- Clickable links
- Totals at the bottom
- Clean professional layout
"""

import json
import sys
import os
from pathlib import Path

# Configure UTF-8 encoding for Windows
if sys.platform == 'win32':
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8')

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo
except ImportError:
    print("❌ Error: Necesitas instalar openpyxl")
    print("💡 Ejecuta: pip install openpyxl")
    sys.exit(1)

import argparse
import glob
import re

# ============================================================================
# CONFIGURATION
# ============================================================================

# Folder for JSON files by department
OUTPUT_DIR = 'OutputDepartamentos'

# Default files (if no department is specified)
JSON_INPUT = 'resultados_validados_gemini.json'
EXCEL_OUTPUT = 'resultados_validados_gemini.xlsx'

# File with missing JRVs
JRVS_FALTANTES_FILE = 'jrvs_faltantes.js'

# Department mapping
DEPARTAMENTOS = {
    "01": "ATLANTIDA",
    "02": "COLON",
    "03": "COMAYAGUA",
    "04": "COPAN",
    "05": "CORTES",
    "06": "CHOLUTECA",
    "07": "EL PARAISO",
    "08": "FRANCISCO MORAZAN",
    "09": "GRACIAS A DIOS",
    "10": "INTIBUCA",
    "11": "ISLAS DE LA BAHIA",
    "12": "LA PAZ",
    "13": "LEMPIRA",
    "14": "OCOTEPEQUE",
    "15": "OLANCHO",
    "16": "SANTA BARBARA",
    "17": "VALLE",
    "18": "YORO",
    "20": "VOTO EN EL EXTERIOR"
}

# Colors (RGB in hexadecimal)
COLOR_HEADER = 'FF2E75B6'         # Professional blue
COLOR_TOTAL = 'FFF4B084'          # Soft orange
COLOR_INCONSISTENCIA = 'FFFFC7CE' # Light red
COLOR_OK = 'FFC6EFCE'             # Light green
COLOR_LINK = 'FF0563C1'           # Blue for links
COLOR_AMARILLO = 'FFFFFF00'       # Yellow for CNE vs AI differences
COLOR_ROJO = 'FFFF0000'           # Red for incorrect sum

# ============================================================================
# FUNCTIONS
# ============================================================================

def cargar_jrvs_faltantes() -> set:
    """Load the set of missing JRVs from jrvs_faltantes.js"""
    print(f"📂 Cargando JRVs faltantes desde: {JRVS_FALTANTES_FILE}")

    if not os.path.exists(JRVS_FALTANTES_FILE):
        print(f"❌ Error: Archivo no encontrado: {JRVS_FALTANTES_FILE}")
        sys.exit(1)

    with open(JRVS_FALTANTES_FILE, 'r', encoding='utf-8') as f:
        contenido = f.read()

    # Extract the JRVS_FALTANTES array using regex
    # Find: const JRVS_FALTANTES = [ ... ];
    match = re.search(r'const\s+JRVS_FALTANTES\s*=\s*\[([\s\S]*?)\];', contenido)

    if not match:
        print(f"❌ Error: No se pudo encontrar JRVS_FALTANTES en {JRVS_FALTANTES_FILE}")
        sys.exit(1)

    # Extract numbers from the array
    array_contenido = match.group(1)
    numeros = re.findall(r'\d+', array_contenido)
    jrvs_faltantes = set(int(num) for num in numeros)

    print(f"✅ Cargados {len(jrvs_faltantes):,} JRVs faltantes\n")
    return jrvs_faltantes

def cargar_json(ruta: str) -> list:
    """Load JSON with validated data"""
    print(f"📂 Cargando datos desde: {ruta}")

    if not os.path.exists(ruta):
        print(f"❌ Error: Archivo no encontrado: {ruta}")
        sys.exit(1)

    with open(ruta, 'r', encoding='utf-8') as f:
        datos = json.load(f)

    if isinstance(datos, dict):
        datos = [datos]

    # Process each record to add calculated fields
    print("🔄 Procesando datos y agregando campos calculados...")
    for registro in datos:
        # Calculate SumCneManual (sum of individual CNE votes)
        sum_cne_manual = (
            registro.get('votos_dc', 0) +
            registro.get('votos_libre', 0) +
            registro.get('votos_pinu', 0) +
            registro.get('votos_liberal', 0) +
            registro.get('votos_nacional', 0) +
            registro.get('votos_blanco', 0) +
            registro.get('votos_nulos', 0)
        )

        # Add SumCneManual field
        registro['SumCneManual'] = sum_cne_manual

        # Check for inconsistency between SumCneManual and pdf_gran_total
        pdf_gran_total = registro.get('pdf_gran_total')
        if pdf_gran_total is not None and sum_cne_manual != pdf_gran_total:
            registro['InconsistenciaGrandTotalManual'] = 1
        else:
            registro['InconsistenciaGrandTotalManual'] = 0

        # ===================================================================
        # NEW: RequiereRevisionActaSumada and MotivoRevision
        # ===================================================================
        motivos = []
        requiere_revision = 0

        # 1. Check InconsistenciaDatosDigitados
        if registro.get('InconsistenciaDatosDigitados', 0) == 1:
            requiere_revision = 1
            motivos.append("Inconsistencia en datos digitados CNE vs IA")

        # 2. Check if manual sum does not match CNE total
        if registro.get('InconsistenciaGrandTotalManual', 0) == 1:
            requiere_revision = 1
            motivos.append(f"Suma manual ({sum_cne_manual}) ≠ Gran Total IA ({pdf_gran_total})")

        # 3. Check ballots: received - unused = used
        pdf_recibidas = registro.get('pdf_papeletas_recibidas')
        pdf_no_utilizadas = registro.get('pdf_papeletas_no_utilizadas')
        pdf_utilizadas = registro.get('pdf_papeletas_utilizadas')

        if pdf_recibidas is not None and pdf_no_utilizadas is not None and pdf_utilizadas is not None:
            if (pdf_recibidas - pdf_no_utilizadas) != pdf_utilizadas:
                requiere_revision = 1
                motivos.append(f"Papeletas: Recibidas({pdf_recibidas}) - No utilizadas({pdf_no_utilizadas}) ≠ Utilizadas({pdf_utilizadas})")

        # 4. Check used ballots = total voters
        pdf_total_votantes = registro.get('pdf_total_votantes')
        if pdf_utilizadas is not None and pdf_total_votantes is not None:
            if pdf_utilizadas != pdf_total_votantes:
                requiere_revision = 1
                motivos.append(f"Papeletas utilizadas({pdf_utilizadas}) ≠ Total votantes({pdf_total_votantes})")

        # Add fields
        registro['RequiereRevisionActaSumada'] = requiere_revision
        registro['MotivoRevision'] = "; ".join(motivos) if motivos else "OK - No requiere revisión"

        # ===================================================================
        # NEW: Fraud detection against Liberal party
        # ===================================================================

        # LiberalVotosFaltantes: CNE < AI (Liberal has fewer votes than it should)
        votos_liberal_cne = registro.get('votos_liberal', 0)
        pdf_votos_liberal = registro.get('pdf_votos_liberal')

        if pdf_votos_liberal is not None and votos_liberal_cne < pdf_votos_liberal:
            registro['LiberalVotosFaltantes'] = 1
        else:
            registro['LiberalVotosFaltantes'] = 0

        # NacionalVotosSobrantes: CNE > AI (Nacional has more votes than it should)
        votos_nacional_cne = registro.get('votos_nacional', 0)
        pdf_votos_nacional = registro.get('pdf_votos_nacional')

        if pdf_votos_nacional is not None and votos_nacional_cne > pdf_votos_nacional:
            registro['NacionalVotosSobrantes'] = 1
        else:
            registro['NacionalVotosSobrantes'] = 0

    print(f"✅ Cargados {len(datos)} registros\n")
    return datos

def es_campo_numerico(valor) -> bool:
    """Check if a field is numeric"""
    return isinstance(valor, (int, float)) and not isinstance(valor, bool)

def es_campo_link(nombre_campo: str) -> bool:
    """Check if a field is a link"""
    return 'url' in nombre_campo.lower() or 'link' in nombre_campo.lower()

def es_campo_inconsistencia(nombre_campo: str) -> bool:
    """Check if a field is an inconsistency field"""
    nombre_lower = nombre_campo.lower()
    return 'inconsistencia' in nombre_lower or nombre_lower.startswith('es_inconsistente')

def aplicar_estilo_header(ws, fila: int, total_columnas: int):
    """Apply style to the header"""
    for col in range(1, total_columnas + 1):
        celda = ws.cell(row=fila, column=col)
        celda.font = Font(bold=True, color='FFFFFFFF', size=11)
        celda.fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type='solid')
        celda.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        celda.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

def aplicar_estilo_total(ws, fila: int, total_columnas: int):
    """Apply style to the totals row"""
    for col in range(1, total_columnas + 1):
        celda = ws.cell(row=fila, column=col)
        celda.font = Font(bold=True, size=11)
        celda.fill = PatternFill(start_color=COLOR_TOTAL, end_color=COLOR_TOTAL, fill_type='solid')
        celda.alignment = Alignment(horizontal='center', vertical='center')
        celda.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='medium'),
            bottom=Side(style='medium')
        )

def aplicar_formato_condicional(ws, fila: int, col: int, valor, nombre_campo: str, registro: dict):
    """Apply conditional formatting based on data type"""
    celda = ws.cell(row=fila, column=col)

    # Borders for all cells
    celda.border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Alignment
    if es_campo_numerico(valor):
        celda.alignment = Alignment(horizontal='right', vertical='center')
    else:
        celda.alignment = Alignment(horizontal='left', vertical='center')

    # YELLOW FORMAT: Differences between CNE and AI
    # Compare votos_X (CNE) with pdf_votos_X (AI)
    campos_comparacion = {
        'votos_dc': 'pdf_votos_dc',
        'votos_libre': 'pdf_votos_libre',
        'votos_pinu': 'pdf_votos_pinu',
        'votos_liberal': 'pdf_votos_liberal',
        'votos_nacional': 'pdf_votos_nacional',
        'votos_blanco': 'pdf_votos_blanco',
        'votos_nulos': 'pdf_votos_nulos',
        'pdf_votos_dc': 'votos_dc',
        'pdf_votos_libre': 'votos_libre',
        'pdf_votos_pinu': 'votos_pinu',
        'pdf_votos_liberal': 'votos_liberal',
        'pdf_votos_nacional': 'votos_nacional',
        'pdf_votos_blanco': 'votos_blanco',
        'pdf_votos_nulos': 'votos_nulos',
    }

    # Check if this field has a CNE vs AI difference
    if nombre_campo in campos_comparacion:
        campo_comparar = campos_comparacion[nombre_campo]
        valor_cne = registro.get(nombre_campo) if nombre_campo.startswith('votos_') else registro.get(campo_comparar)
        valor_ia = registro.get(campo_comparar) if nombre_campo.startswith('votos_') else registro.get(nombre_campo)

        if valor_cne != valor_ia and valor_ia is not None:
            celda.fill = PatternFill(start_color=COLOR_AMARILLO, end_color=COLOR_AMARILLO, fill_type='solid')
            celda.font = Font(bold=True)
            return

    # Compare GrandTotal (SumCneManual vs pdf_gran_total)
    if nombre_campo == 'SumCneManual' or nombre_campo == 'pdf_gran_total':
        sum_cne = registro.get('SumCneManual')
        pdf_total = registro.get('pdf_gran_total')

        if sum_cne != pdf_total and pdf_total is not None:
            celda.fill = PatternFill(start_color=COLOR_AMARILLO, end_color=COLOR_AMARILLO, fill_type='solid')
            celda.font = Font(bold=True)
            return

    # RED FORMAT: SumCneManual when it doesn't match pdf_gran_total
    if nombre_campo == 'SumCneManual' and registro.get('InconsistenciaGrandTotalManual', 0) == 1:
        celda.fill = PatternFill(start_color=COLOR_ROJO, end_color=COLOR_ROJO, fill_type='solid')
        celda.font = Font(bold=True, color='FFFFFFFF')
        return

    # RED FORMAT: InconsistenciaGrandTotalManual when it is 1
    if nombre_campo == 'InconsistenciaGrandTotalManual' and valor == 1:
        celda.fill = PatternFill(start_color=COLOR_ROJO, end_color=COLOR_ROJO, fill_type='solid')
        celda.font = Font(bold=True, color='FFFFFFFF')
        return

    # RED FORMAT: RequiereRevisionActaSumada when it is 1
    if nombre_campo == 'RequiereRevisionActaSumada' and valor == 1:
        celda.fill = PatternFill(start_color=COLOR_ROJO, end_color=COLOR_ROJO, fill_type='solid')
        celda.font = Font(bold=True, color='FFFFFFFF')
        return

    # YELLOW FORMAT: MotivoRevision when it is not "OK"
    if nombre_campo == 'MotivoRevision' and valor != "OK - No requiere revisión":
        celda.fill = PatternFill(start_color=COLOR_AMARILLO, end_color=COLOR_AMARILLO, fill_type='solid')
        celda.font = Font(bold=True)
        celda.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        return

    # RED FORMAT: LiberalVotosFaltantes when it is 1 (fraud)
    if nombre_campo == 'LiberalVotosFaltantes' and valor == 1:
        celda.fill = PatternFill(start_color=COLOR_ROJO, end_color=COLOR_ROJO, fill_type='solid')
        celda.font = Font(bold=True, color='FFFFFFFF')
        return

    # RED FORMAT: NacionalVotosSobrantes when it is 1 (fraud)
    if nombre_campo == 'NacionalVotosSobrantes' and valor == 1:
        celda.fill = PatternFill(start_color=COLOR_ROJO, end_color=COLOR_ROJO, fill_type='solid')
        celda.font = Font(bold=True, color='FFFFFFFF')
        return

    # Background color for inconsistencies
    if es_campo_inconsistencia(nombre_campo) and valor == 1:
        celda.fill = PatternFill(start_color=COLOR_INCONSISTENCIA, end_color=COLOR_INCONSISTENCIA, fill_type='solid')
        celda.font = Font(bold=True)
    elif nombre_campo == 'SumatoriaManualPorPartido' or nombre_campo == 'pdf_gran_total':
        celda.font = Font(bold=True)

def autoajustar_columnas(ws, datos: list, headers: list):
    """Auto-fit column widths based on content"""
    for idx, header in enumerate(headers, 1):
        max_length = len(str(header))

        # Check content of all rows
        for dato in datos:
            valor = dato.get(header, '')
            if valor is not None:
                valor_str = str(valor)
                # Limit max width to 80 characters
                if es_campo_link(header):
                    # For links, use shorter "Ver acta" text
                    max_length = max(max_length, len("Ver acta"))
                elif header == 'MotivoRevision':
                    # For MotivoRevision, use larger width (up to 100 characters)
                    max_length = max(max_length, min(len(valor_str), 100))
                else:
                    max_length = max(max_length, min(len(valor_str), 80))

        # Set column width (add padding)
        adjusted_width = max_length + 2
        ws.column_dimensions[get_column_letter(idx)].width = adjusted_width

def crear_excel(datos: list, archivo_salida: str):
    """Create Excel file with professional formatting"""
    print("📊 Creando archivo Excel...")

    if len(datos) == 0:
        print("⚠️  No hay datos para exportar")
        return

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Resultados Validados"

    # Get headers (all keys from first record)
    headers_originales = list(datos[0].keys())

    # Reorder headers: numero_jrv and url_drive first, RequiereRevisionActaSumada and MotivoRevision last
    headers = []

    # Columns at the start
    columnas_inicio = ['numero_jrv', 'url_drive']

    # Columns at the end
    columnas_final = ['RequiereRevisionActaSumada', 'MotivoRevision', 'LiberalVotosFaltantes', 'NacionalVotosSobrantes']

    # Add starting columns
    for col in columnas_inicio:
        if col in headers_originales:
            headers.append(col)

    # Add remaining columns (excluding start and end columns)
    for h in headers_originales:
        if h not in columnas_inicio and h not in columnas_final:
            headers.append(h)

    # Add ending columns
    for col in columnas_final:
        if col in headers_originales:
            headers.append(col)

    total_columnas = len(headers)

    # Identify numeric columns for summing
    columnas_numericas = {}
    for idx, header in enumerate(headers, 1):
        primer_valor = datos[0].get(header)
        if es_campo_numerico(primer_valor):
            columnas_numericas[idx] = header

    # Write headers
    print("✍️  Escribiendo encabezados...")
    for idx, header in enumerate(headers, 1):
        ws.cell(row=1, column=idx, value=header)

    # Write data
    print(f"✍️  Escribiendo {len(datos)} filas de datos...")
    for fila_idx, dato in enumerate(datos, 2):
        for col_idx, header in enumerate(headers, 1):
            valor = dato.get(header, '')

            # Convert None to empty string
            if valor is None:
                valor = ''

            # Write the value
            ws.cell(row=fila_idx, column=col_idx, value=valor)

            # If it's a link, convert it to a hyperlink
            if es_campo_link(header) and valor and str(valor).startswith('http'):
                celda = ws.cell(row=fila_idx, column=col_idx)
                celda.hyperlink = str(valor)
                celda.font = Font(color=COLOR_LINK, underline='single')
                celda.value = "Ver acta"  # Shorter and cleaner text

            # Apply conditional formatting (pass the full record)
            aplicar_formato_condicional(ws, fila_idx, col_idx, valor, header, dato)

    # Apply table format to data (not including totals)
    print("📋 Aplicando formato de tabla con filtros...")
    ultima_fila_datos = len(datos) + 1
    ultima_columna = get_column_letter(total_columnas)

    # Create Excel table (includes headers and data, NOT totals)
    tabla_rango = f"A1:{ultima_columna}{ultima_fila_datos}"
    tabla = Table(displayName="TablaResultados", ref=tabla_rango)

    # Table style (medium blue with stripes)
    estilo_tabla = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    tabla.tableStyleInfo = estilo_tabla
    ws.add_table(tabla)

    # Leave a blank row as separator
    fila_separador = len(datos) + 2

    # Add TOTALS row (after the blank row)
    print("🔢 Calculando totales...")
    fila_total = len(datos) + 3

    # First column: "TOTAL" text
    ws.cell(row=fila_total, column=1, value="TOTAL")

    # Sum numeric columns (excluding separator row)
    for col_idx, nombre_campo in columnas_numericas.items():
        # Use SUM formula to calculate total (data only, not separator)
        letra_col = get_column_letter(col_idx)
        formula = f"=SUM({letra_col}2:{letra_col}{ultima_fila_datos})"
        ws.cell(row=fila_total, column=col_idx, value=formula)

    # Apply style to the totals row
    aplicar_estilo_total(ws, fila_total, total_columnas)

    # Auto-fit column widths
    print("📏 Autoajustando ancho de columnas...")
    autoajustar_columnas(ws, datos, headers)

    # Freeze panes: First row + numero_jrv and url_drive columns
    # Find key columns to freeze
    columnas_clave = ['numero_jrv', 'url_drive']
    ultima_col_fija = 0

    for col_clave in columnas_clave:
        if col_clave in headers:
            idx = headers.index(col_clave) + 1
            ultima_col_fija = max(ultima_col_fija, idx)

    # If key columns found, freeze up to and including the last one
    if ultima_col_fija > 0:
        letra_col = get_column_letter(ultima_col_fija + 1)
        ws.freeze_panes = f'{letra_col}2'  # Freeze columns up to ultima_col_fija and row 1
        columnas_fijadas = [headers[i] for i in range(ultima_col_fija)]
        print(f"🔒 Columnas fijadas: 1-{ultima_col_fija} ({', '.join(columnas_fijadas)})")
    else:
        # If no key columns, only freeze the header
        ws.freeze_panes = 'A2'
        print(f"🔒 Solo encabezado fijado")

    # Save file
    print(f"💾 Guardando archivo: {archivo_salida}")
    wb.save(archivo_salida)
    print(f"✅ Archivo Excel creado exitosamente!\n")

def mostrar_resumen(datos: list):
    """Show a summary of statistics"""
    print("=" * 80)
    print("📊 RESUMEN DE DATOS")
    print("=" * 80)
    print(f"Total de registros: {len(datos)}")

    # Count inconsistencies
    inconsistencias = sum(1 for d in datos if d.get('InconsistenciaDatosDigitados', 0) == 1)
    print(f"Actas con inconsistencias: {inconsistencias}")
    print(f"Actas correctas: {len(datos) - inconsistencias}")

    if len(datos) > 0:
        porcentaje = ((len(datos) - inconsistencias) / len(datos)) * 100
        print(f"Porcentaje de precisión: {porcentaje:.1f}%")

    # Sum votes
    total_nacional = sum(d.get('pdf_votos_nacional', 0) or 0 for d in datos)
    total_liberal = sum(d.get('pdf_votos_liberal', 0) or 0 for d in datos)
    total_libre = sum(d.get('pdf_votos_libre', 0) or 0 for d in datos)

    print(f"\n🗳️  TOTALES DE VOTOS (desde PDF):")
    print(f"   Partido Nacional: {total_nacional:,}")
    print(f"   Partido Liberal: {total_liberal:,}")
    print(f"   Partido LIBRE: {total_libre:,}")
    print("=" * 80)
    print()

def main():
    # Parse command-line arguments
    parser = argparse.ArgumentParser(
        description='Conversor JSON a Excel - Resultados Validados',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog='''
Ejemplos de uso:
  python json_to_excel_formateado.py --depto 05          # Solo CORTES
  python json_to_excel_formateado.py --depto 08          # Solo FRANCISCO MORAZAN
  python json_to_excel_formateado.py                      # Todos los departamentos (consolidado)
  python json_to_excel_formateado.py --solofaltantes 1   # Solo JRVs faltantes

Códigos de departamentos:
  01 = ATLANTIDA          08 = FRANCISCO MORAZAN    15 = OLANCHO
  02 = COLON              09 = GRACIAS A DIOS       16 = SANTA BARBARA
  03 = COMAYAGUA          10 = INTIBUCA             17 = VALLE
  04 = COPAN              11 = ISLAS DE LA BAHIA    18 = YORO
  05 = CORTES             12 = LA PAZ               20 = VOTO EN EL EXTERIOR
  06 = CHOLUTECA          13 = LEMPIRA
  07 = EL PARAISO         14 = OCOTEPEQUE
        '''
    )
    parser.add_argument('--depto', type=str, help='Código del departamento (ej: 05 para CORTES)')
    parser.add_argument('--solofaltantes', type=int, default=0, choices=[0, 1],
                        help='Si es 1, solo convierte JRVs que existen en jrvs_faltantes.js (default: 0)')
    args = parser.parse_args()

    print("=" * 80)
    print("CONVERSOR JSON A EXCEL - RESULTADOS VALIDADOS")
    print("=" * 80)
    print()

    # Show filter info if enabled
    if args.solofaltantes == 1:
        print("⚠️  FILTRO ACTIVO: Solo se procesarán JRVs faltantes")
        print(f"📋 Archivo de referencia: {JRVS_FALTANTES_FILE}\n")

    # Determine input and output files based on parameter
    archivo_entrada = JSON_INPUT
    archivo_salida = EXCEL_OUTPUT

    if args.depto:
        # Process specific department
        id_departamento = args.depto.zfill(2)  # Ensure 2 digits

        if id_departamento not in DEPARTAMENTOS:
            print(f"❌ Error: Departamento '{id_departamento}' no válido")
            print("\n📋 Departamentos disponibles:")
            for id_dep, nombre in sorted(DEPARTAMENTOS.items()):
                print(f"   {id_dep} = {nombre}")
            sys.exit(1)

        nombre_departamento = DEPARTAMENTOS[id_departamento]
        archivo_entrada = os.path.join(OUTPUT_DIR, f"{id_departamento}_{nombre_departamento.replace(' ', '_')}.json")
        archivo_salida = f"{id_departamento}_{nombre_departamento.replace(' ', '_')}.xlsx"

        print(f"🎯 Departamento seleccionado: {id_departamento} - {nombre_departamento}")
        print(f"📂 Archivo JSON: {archivo_entrada}")
        print(f"📊 Archivo Excel: {archivo_salida}\n")
    else:
        # Consolidate all departments
        print(f"📋 Consolidando TODOS los departamentos...")
        print(f"📂 Buscando archivos en: {OUTPUT_DIR}/\n")

        # Find all JSON files in OutputDepartamentos
        patron = os.path.join(OUTPUT_DIR, "*.json")
        archivos_json = glob.glob(patron)

        if not archivos_json:
            print(f"⚠️  No se encontraron archivos JSON en {OUTPUT_DIR}/")
            print(f"   Usando archivo por defecto: {archivo_entrada}\n")
        else:
            # Consolidate all JSONs
            print(f"✅ Encontrados {len(archivos_json)} archivos JSON")
            datos_consolidados = []

            for archivo in sorted(archivos_json):
                try:
                    with open(archivo, 'r', encoding='utf-8') as f:
                        datos_depto = json.load(f)
                        if isinstance(datos_depto, list):
                            datos_consolidados.extend(datos_depto)
                            print(f"   ✅ {os.path.basename(archivo)}: {len(datos_depto)} actas")
                except Exception as e:
                    print(f"   ⚠️  Error en {os.path.basename(archivo)}: {e}")

            if datos_consolidados:
                # Temporarily save consolidated data
                archivo_temp = "temp_consolidado.json"
                with open(archivo_temp, 'w', encoding='utf-8') as f:
                    json.dump(datos_consolidados, f, ensure_ascii=False, indent=2)

                archivo_entrada = archivo_temp
                print(f"\n📊 Total consolidado: {len(datos_consolidados)} actas\n")
            else:
                print(f"\n⚠️  No se pudo consolidar datos, usando archivo por defecto\n")

    # Load data
    datos = cargar_json(archivo_entrada)

    if len(datos) == 0:
        print("⚠️  El archivo JSON está vacío. No se generará Excel.")
        return

    # Filter only missing JRVs if solofaltantes=1
    if args.solofaltantes == 1:
        print("🔍 Aplicando filtro: Solo JRVs faltantes")
        jrvs_faltantes = cargar_jrvs_faltantes()

        total_antes = len(datos)
        datos = [d for d in datos if d.get('numero_jrv') in jrvs_faltantes]
        total_despues = len(datos)

        print(f"📊 Registros antes del filtro: {total_antes:,}")
        print(f"📊 Registros después del filtro: {total_despues:,}")
        print(f"📊 Registros filtrados: {total_antes - total_despues:,}\n")

        if total_despues == 0:
            print("⚠️  No hay JRVs faltantes en el archivo JSON. No se generará Excel.")
            return

        # Modify output filename to include "_faltantes"
        base_name, ext = os.path.splitext(archivo_salida)
        archivo_salida = f"{base_name}_faltantes{ext}"
        print(f"📝 Archivo de salida modificado: {archivo_salida}\n")

    # Sort data by numero_jrv (ascending)
    print("🔢 Ordenando datos por numero_jrv...")
    datos.sort(key=lambda x: x.get('numero_jrv', 0))
    print(f"✅ Datos ordenados: JRV {datos[0].get('numero_jrv', 'N/A')} → {datos[-1].get('numero_jrv', 'N/A')}\n")

    # Show summary
    mostrar_resumen(datos)

    # Create Excel
    crear_excel(datos, archivo_salida)

    # Clean up temp file if it exists
    if 'archivo_temp' in locals() and os.path.exists(archivo_temp):
        os.remove(archivo_temp)

    # Final message
    print("=" * 80)
    print("✅ CONVERSIÓN COMPLETADA")
    print("=" * 80)
    print(f"📁 Archivo generado: {archivo_salida}")
    print(f"📊 Total de filas: {len(datos)}")
    print(f"📏 Total de columnas: {len(datos[0].keys()) if datos else 0}")
    print()
    print("💡 Características del archivo:")
    print("   ✅ Formato de tabla de Excel con filtros automáticos")
    print("   ✅ Bandas de colores para mejor lectura")
    print("   ✅ Columnas autoajustadas al contenido")
    print("   ✅ Links clicleables (columna url_drive)")
    print("   ✅ Fila de totales con sumatorias (separada por fila en blanco)")
    print("   ✅ Formato condicional para inconsistencias")
    print("   ✅ Encabezados congelados")
    print("=" * 80)

if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\n\n⚠️  Proceso interrumpido por el usuario")
    except Exception as e:
        print(f"\n❌ Error inesperado: {e}")
        import traceback
        traceback.print_exc()
